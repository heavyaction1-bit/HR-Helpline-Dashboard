from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import warnings
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook


REQUIRED_SHEETS = ("SR Data", "Sickness", "Annual leave")
PLACEHOLDER_HEADER_RE = re.compile(r"^(column|unnamed)\s*:?\s*\d+$", re.IGNORECASE)


@dataclass
class LoadedWorkbook:
    sr_data: pd.DataFrame
    sickness: pd.DataFrame
    annual_leave: pd.DataFrame
    stats: dict[str, dict[str, int | str | list[str]]]
    missing_sheets: list[str]


def _source_to_bytes(source: str | Path | bytes | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if hasattr(source, "getvalue"):
        value = source.getvalue()
        return value if isinstance(value, bytes) else bytes(value)
    if hasattr(source, "read"):
        current_position = None
        if hasattr(source, "tell") and hasattr(source, "seek"):
            current_position = source.tell()
            source.seek(0)
        data = source.read()
        if current_position is not None:
            source.seek(current_position)
        return data
    raise TypeError("Unsupported workbook source. Provide a path, bytes, or file-like object.")


def _is_placeholder_header(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return not text or bool(PLACEHOLDER_HEADER_RE.match(text))


def _meaningful_header_indexes(workbook_bytes: bytes, sheet_name: str) -> list[int]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        indexes = [
            index
            for index, value in enumerate(header_row)
            if not _is_placeholder_header(value)
        ]
        if indexes:
            return indexes
        return [index for index, value in enumerate(header_row) if value is not None]
    finally:
        workbook.close()


def read_excel_sheet(workbook_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    usecols = _meaningful_header_indexes(workbook_bytes, sheet_name)
    if not usecols:
        return pd.DataFrame()

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        dataframe = pd.read_excel(
            BytesIO(workbook_bytes),
            sheet_name=sheet_name,
            usecols=usecols,
            engine="openpyxl",
        )

    dataframe = dataframe.dropna(how="all")
    return dataframe


def load_raw_workbook(source: str | Path | bytes | BinaryIO) -> LoadedWorkbook:
    workbook_bytes = _source_to_bytes(source)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        excel_file = pd.ExcelFile(BytesIO(workbook_bytes), engine="openpyxl")

    sheet_names = set(excel_file.sheet_names)
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheet_names]
    if missing_sheets:
        missing = ", ".join(missing_sheets)
        raise ValueError(f"Workbook is missing required sheet(s): {missing}")

    raw_sheets = {
        sheet_name: read_excel_sheet(workbook_bytes, sheet_name)
        for sheet_name in REQUIRED_SHEETS
    }
    stats = {
        sheet_name: {
            "rows_loaded": int(dataframe.shape[0]),
            "columns_loaded": int(dataframe.shape[1]),
            "columns": [str(column) for column in dataframe.columns],
        }
        for sheet_name, dataframe in raw_sheets.items()
    }

    return LoadedWorkbook(
        sr_data=raw_sheets["SR Data"],
        sickness=raw_sheets["Sickness"],
        annual_leave=raw_sheets["Annual leave"],
        stats=stats,
        missing_sheets=missing_sheets,
    )
